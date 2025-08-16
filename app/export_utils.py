import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

def export_to_excel(
    org_name,
    cms_data=None,
    google_reviews=None,
    yelp_reviews=None,
    about_data=None,
    other_data=None,
    export_dir="exports"
):
    """
    Export all collected data to an Excel workbook with separate sheets.
    Includes CMS, Google Reviews, Yelp Reviews, About, and manual data.
    """
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = org_name.replace(" ", "_").replace("/", "_")
    filename = os.path.join(export_dir, f"{safe_name}_{timestamp}.xlsx")
    
    wb = Workbook()
    
    # --- CMS Sheet ---
    if cms_data is not None:
        ws = wb.active
        ws.title = "CMS Data"
        if hasattr(cms_data, "to_dict"):
            cms_dict = cms_data.to_dict()
        else:
            cms_dict = cms_data
        for r_idx, (key, value) in enumerate(cms_dict.items(), start=1):
            ws.cell(row=r_idx, column=1, value=key)
            ws.cell(row=r_idx, column=2, value=value)
    
    # --- Google Reviews Sheet ---
    if google_reviews:
        ws = wb.create_sheet(title="Google Reviews")
        df_google = pd.DataFrame(google_reviews)
        for r in dataframe_to_rows(df_google, index=False, header=True):
            ws.append(r)
    
    # --- Yelp Reviews Sheet ---
    if yelp_reviews:
        ws = wb.create_sheet(title="Yelp Reviews")
        df_yelp = pd.DataFrame(yelp_reviews)
        for r in dataframe_to_rows(df_yelp, index=False, header=True):
            ws.append(r)
    
    # --- About Data Sheet ---
    if about_data:
        ws = wb.create_sheet(title="About Data")
        for key, value in (about_data.items() if isinstance(about_data, dict) else {}).items():
            ws.append([key, value])
    
    # --- Manual Data Sheets ---
    if other_data:
        # US News
        usnews = other_data.get("usnews")
        if usnews:
            ws = wb.create_sheet(title="US News Manual")
            for key, value in usnews.items():
                ws.append([key, value])
        
        # Manual Yelp
        manual_yelp = other_data.get("yelp")
        if manual_yelp:
            ws = wb.create_sheet(title="Yelp Manual")
            df_manual_yelp = pd.DataFrame(manual_yelp)
            for r in dataframe_to_rows(df_manual_yelp, index=False, header=True):
                ws.append(r)
        
        # Other freepaste
        other = other_data.get("other")
        if other:
            ws = wb.create_sheet(title="Other Data")
            for key, value in other.items():
                ws.append([key, value])
    
    wb.save(filename)
    return filename
